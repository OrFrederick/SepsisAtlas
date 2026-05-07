"""Batch-parse PDFs in ``data/papers/raw/`` into JSON + upsert ``papers`` rows.

Usage::

    python -m parse.run_parse                 # parse all, skip already-parsed
    python -m parse.run_parse --force         # re-parse everything
    python -m parse.run_parse --jobs 2        # override worker count (default 4)
    python -m parse.run_parse --only Gai_2022.pdf Seymour_2016.pdf
"""

from __future__ import annotations

import argparse
import hashlib
import json
import multiprocessing as mp
import os
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

from sepsis_atlas.config import PAPERS_RAW, PAPERS_PARSED, DATA_DIR, PIPELINE_VERSION
from sepsis_atlas.db import Paper, get_engine, init_db


PARSER_VERSION_FAILED = "failed"
DEFAULT_WORKERS = 4
DOI_INDEX_PATH = DATA_DIR / "papers" / "_index.xlsx"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def load_doi_index(xlsx_path: Path) -> dict[str, str]:
    """Return ``{file_name: doi}`` from the organizer index."""
    if not xlsx_path.exists():
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        return {}

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def find_col(*candidates: str) -> int | None:
        for cand in candidates:
            for i, h in enumerate(header):
                if cand in h:
                    return i
        return None

    fn_col = find_col("file_name", "filename", "file")
    doi_col = find_col("doi")
    if fn_col is None or doi_col is None:
        return {}

    out: dict[str, str] = {}
    for row in rows[1:]:
        if not row or fn_col >= len(row) or doi_col >= len(row):
            continue
        fn, doi = row[fn_col], row[doi_col]
        if not fn or not doi:
            continue
        fn = str(fn).strip()
        if not fn.lower().endswith(".pdf"):
            fn = f"{fn}.pdf"
        out[fn] = str(doi).strip()
    return out


def _worker_parse(pdf_path_str: str) -> dict[str, Any]:
    """Run in a worker process; returns a small status dict."""
    # Lazy import so the heavy docling import only happens inside workers.
    from parse.docling_parser import parse_pdf, PARSER_VERSION

    pdf_path = Path(pdf_path_str)
    out_path = PAPERS_PARSED / f"{pdf_path.stem}.json"
    t0 = time.time()
    try:
        parsed = parse_pdf(pdf_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w") as f:
            json.dump(parsed, f, ensure_ascii=False)
        return {
            "file_name": pdf_path.name,
            "ok": True,
            "title": parsed.get("title"),
            "n_sections": len(parsed.get("sections", [])),
            "n_tables": len(parsed.get("tables", [])),
            "parser_version": PARSER_VERSION,
            "elapsed_s": time.time() - t0,
            "out_path": str(out_path),
        }
    except Exception as e:  # noqa: BLE001
        return {
            "file_name": pdf_path.name,
            "ok": False,
            "error": f"{type(e).__name__}: {e}",
            "traceback": traceback.format_exc(),
            "parser_version": PARSER_VERSION_FAILED,
            "elapsed_s": time.time() - t0,
        }


def _silence_workers():
    """Reduce log spam in worker processes."""
    import logging

    logging.getLogger().setLevel(logging.WARNING)
    for name in ("docling", "docling_core", "RapidOCR", "transformers"):
        logging.getLogger(name).setLevel(logging.WARNING)
    os.environ.setdefault("HF_HUB_DISABLE_PROGRESS_BARS", "1")
    os.environ.setdefault("TRANSFORMERS_VERBOSITY", "error")


def _upsert_paper(
    engine,
    *,
    file_name: str,
    pdf_path: Path,
    doi: str | None,
    title: str | None,
    pdf_hash: str,
    parser_version: str,
    run_id: str,
) -> None:
    stmt = sqlite_insert(Paper).values(
        file_name=file_name,
        doi=doi,
        title=title,
        pdf_hash=pdf_hash,
        parser_version=parser_version,
        source="provided",
        ingest_ts=datetime.utcnow(),
        run_id=run_id,
        pipeline_version=PIPELINE_VERSION,
    )
    update_cols = {
        "doi": stmt.excluded.doi,
        "title": stmt.excluded.title,
        "pdf_hash": stmt.excluded.pdf_hash,
        "parser_version": stmt.excluded.parser_version,
        "source": stmt.excluded.source,
        "ingest_ts": stmt.excluded.ingest_ts,
        "run_id": stmt.excluded.run_id,
        "pipeline_version": stmt.excluded.pipeline_version,
    }
    stmt = stmt.on_conflict_do_update(index_elements=[Paper.file_name], set_=update_cols)
    with engine.begin() as conn:
        conn.execute(stmt)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Parse PDFs with Docling.")
    ap.add_argument("--force", action="store_true", help="Re-parse already-parsed PDFs.")
    ap.add_argument("--jobs", type=int, default=DEFAULT_WORKERS, help="Parallel workers.")
    ap.add_argument("--only", nargs="+", default=None, help="Only parse these filenames.")
    ap.add_argument(
        "--no-db",
        action="store_true",
        help="Skip writing rows to the papers table (filesystem only).",
    )
    args = ap.parse_args(argv)

    _silence_workers()

    if not PAPERS_RAW.exists():
        print(f"ERROR: {PAPERS_RAW} does not exist.", file=sys.stderr)
        return 2

    pdfs = sorted(PAPERS_RAW.glob("*.pdf"))
    if args.only:
        wanted = {Path(name).name for name in args.only}
        pdfs = [p for p in pdfs if p.name in wanted]
    if not pdfs:
        print("No PDFs to parse.", file=sys.stderr)
        return 1

    PAPERS_PARSED.mkdir(parents=True, exist_ok=True)
    doi_map = load_doi_index(DOI_INDEX_PATH)
    print(f"Loaded DOI index: {len(doi_map)} entries", file=sys.stderr)

    todo: list[Path] = []
    skipped: list[Path] = []
    for p in pdfs:
        out = PAPERS_PARSED / f"{p.stem}.json"
        if out.exists() and not args.force:
            skipped.append(p)
        else:
            todo.append(p)

    print(
        f"To parse: {len(todo)}    Skipped (already parsed): {len(skipped)}    "
        f"Workers: {args.jobs}",
        file=sys.stderr,
    )

    engine = None
    if not args.no_db:
        engine = init_db()

    run_id = datetime.utcnow().strftime("parse-%Y%m%dT%H%M%S")

    results: list[dict[str, Any]] = []

    if todo:
        ctx = mp.get_context("spawn")
        with ctx.Pool(processes=max(1, args.jobs), initializer=_silence_workers) as pool:
            for i, res in enumerate(
                pool.imap_unordered(_worker_parse, [str(p) for p in todo]), 1
            ):
                results.append(res)
                tag = "OK " if res["ok"] else "FAIL"
                detail = (
                    f"sections={res.get('n_sections')} tables={res.get('n_tables')}"
                    if res["ok"]
                    else res.get("error", "")
                )
                print(
                    f"[{i}/{len(todo)}] {tag} {res['file_name']:<28} "
                    f"{res['elapsed_s']:6.1f}s  {detail}",
                    file=sys.stderr,
                    flush=True,
                )

    # Process skipped: still need DB rows.
    for p in skipped:
        out_path = PAPERS_PARSED / f"{p.stem}.json"
        try:
            with out_path.open() as f:
                parsed = json.load(f)
            results.append(
                {
                    "file_name": p.name,
                    "ok": True,
                    "title": parsed.get("title"),
                    "n_sections": len(parsed.get("sections", [])),
                    "n_tables": len(parsed.get("tables", [])),
                    "parser_version": parsed.get("parser", "docling-unknown"),
                    "elapsed_s": 0.0,
                    "skipped": True,
                }
            )
        except Exception as e:  # noqa: BLE001
            results.append(
                {
                    "file_name": p.name,
                    "ok": False,
                    "error": f"reload-skip-failed: {e}",
                    "parser_version": PARSER_VERSION_FAILED,
                    "elapsed_s": 0.0,
                }
            )

    # Upsert papers rows.
    if engine is not None:
        for res in results:
            fname = res["file_name"]
            pdf_path = PAPERS_RAW / fname
            try:
                pdf_hash = sha256_file(pdf_path)
            except Exception:  # noqa: BLE001
                pdf_hash = ""
            _upsert_paper(
                engine,
                file_name=fname,
                pdf_path=pdf_path,
                doi=doi_map.get(fname),
                title=res.get("title"),
                pdf_hash=pdf_hash,
                parser_version=res.get("parser_version") or PARSER_VERSION_FAILED,
                run_id=run_id,
            )

    succ = [r for r in results if r["ok"]]
    fail = [r for r in results if not r["ok"]]
    parsed_now = [r for r in results if r["ok"] and not r.get("skipped")]
    mean_s = (sum(r["elapsed_s"] for r in parsed_now) / len(parsed_now)) if parsed_now else 0.0

    print(
        f"\nDone. success={len(succ)}/{len(results)}  failed={len(fail)}  "
        f"mean_parse_s={mean_s:.1f}",
        file=sys.stderr,
    )
    if fail:
        print("Failures:", file=sys.stderr)
        for r in fail:
            print(f"  - {r['file_name']}: {r.get('error')}", file=sys.stderr)

    return 0 if not fail else 1


if __name__ == "__main__":
    raise SystemExit(main())
