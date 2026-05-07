"""Mechanical structure layer: parsed markdown -> Neo4j Paper/Section/PaperTable/Figure/Reference."""

from __future__ import annotations

import csv
import io
import json
import re
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from api.backends.kg_store import KGStore


_HEADING_RE = re.compile(r"^(#{1,4})\s+(.+?)\s*$")
_PAGE_COMMENT_RE = re.compile(r"<!--\s*page:\s*(\d+)\s*-->")
_DOI_RE = re.compile(r"(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)")
_IMAGE_RE = re.compile(r"!\[(?P<alt>[^\]]*)\]\((?P<src>[^)]*)\)")
_HTML_FIGURE_RE = re.compile(r"<figure[^>]*>(.*?)</figure>", re.DOTALL | re.IGNORECASE)
_REFERENCES_HEADING_RE = re.compile(r"^(references|bibliography)\b", re.IGNORECASE)
_NUMBERED_LIST_RE = re.compile(r"^\s*(?:\d+[.)]|[-*])\s+(.*)$")
_FIGURE_CAPTION_RE = re.compile(r"^\s*(figure|fig\.?)\s*\d+\b", re.IGNORECASE)


def _split_table_block(lines: list[str], start: int) -> tuple[list[str], int]:
    block: list[str] = []
    i = start
    while i < len(lines) and lines[i].lstrip().startswith("|"):
        block.append(lines[i])
        i += 1
    return block, i


def _is_table_separator(line: str) -> bool:
    stripped = line.strip().strip("|").strip()
    if not stripped:
        return False
    cells = [c.strip() for c in stripped.split("|")]
    return all(re.fullmatch(r":?-{3,}:?", c) for c in cells if c)


def _markdown_table_to_csv(block: list[str]) -> str:
    rows: list[list[str]] = []
    for line in block:
        stripped = line.strip()
        if stripped.startswith("|"):
            stripped = stripped[1:]
        if stripped.endswith("|"):
            stripped = stripped[:-1]
        cells = [c.strip() for c in stripped.split("|")]
        rows.append(cells)
    rows = [r for r in rows if not (r and all(re.fullmatch(r":?-{3,}:?", c or "") for c in r))]
    out = io.StringIO()
    writer = csv.writer(out)
    for r in rows:
        writer.writerow(r)
    return out.getvalue().rstrip("\n")


def _last_page_before(text_lines: list[str], idx: int) -> int | None:
    for j in range(idx - 1, -1, -1):
        m = _PAGE_COMMENT_RE.search(text_lines[j])
        if m:
            return int(m.group(1))
    return None


def _section_page(body: str) -> int | None:
    m = _PAGE_COMMENT_RE.search(body)
    return int(m.group(1)) if m else None


def _parse_markdown(md_text: str) -> dict[str, list[dict[str, Any]]]:
    lines = md_text.splitlines()
    sections: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    figures: list[dict[str, Any]] = []
    references: list[dict[str, Any]] = []

    section_starts: list[tuple[int, int, str]] = []
    for i, line in enumerate(lines):
        m = _HEADING_RE.match(line)
        if m:
            section_starts.append((i, len(m.group(1)), m.group(2).strip()))

    refs_section_idx: int | None = None
    for idx, (_, _, heading) in enumerate(section_starts):
        if _REFERENCES_HEADING_RE.match(heading):
            refs_section_idx = idx
            break

    for idx, (line_i, level, heading) in enumerate(section_starts):
        end = section_starts[idx + 1][0] if idx + 1 < len(section_starts) else len(lines)
        body = "\n".join(lines[line_i + 1 : end]).strip()
        if not body and not heading:
            continue
        sections.append(
            {
                "heading": heading,
                "level": level,
                "page": _section_page(body),
                "text": body,
            }
        )

    i = 0
    while i < len(lines):
        line = lines[i]
        if (
            line.lstrip().startswith("|")
            and i + 1 < len(lines)
            and _is_table_separator(lines[i + 1])
        ):
            block, next_i = _split_table_block(lines, i)
            page = _last_page_before(lines, i)
            caption = ""
            for j in range(i - 1, max(-1, i - 5), -1):
                prev = lines[j].strip()
                if not prev:
                    continue
                if _PAGE_COMMENT_RE.search(prev) or _HEADING_RE.match(prev):
                    break
                caption = prev
                break
            tables.append(
                {
                    "page": page,
                    "caption": caption,
                    "csv": _markdown_table_to_csv(block),
                }
            )
            i = next_i
            continue
        i += 1

    for i, line in enumerate(lines):
        for m in _IMAGE_RE.finditer(line):
            page = _last_page_before(lines, i)
            alt = m.group("alt").strip()
            caption = alt
            for j in range(i + 1, min(len(lines), i + 4)):
                nxt = lines[j].strip()
                if not nxt:
                    continue
                if _FIGURE_CAPTION_RE.match(nxt):
                    caption = nxt
                break
            figures.append({"page": page, "caption": caption})
    for m in _HTML_FIGURE_RE.finditer(md_text):
        figures.append({"page": None, "caption": m.group(1).strip()[:500]})

    if refs_section_idx is not None:
        ref_line_i = section_starts[refs_section_idx][0]
        next_heading_end = (
            section_starts[refs_section_idx + 1][0]
            if refs_section_idx + 1 < len(section_starts)
            else len(lines)
        )
        # Trailing tables (markdown pipe-tables) often follow the references
        # block without an intervening heading — Docling renders them flat.
        # Cut the ref block at the first table-row marker so we don't ingest
        # table content as bogus citations.
        ref_end = next_heading_end
        for j in range(ref_line_i + 1, next_heading_end):
            stripped = lines[j].lstrip()
            if stripped.startswith("|") or stripped.startswith("|---"):
                ref_end = j
                break
        ref_lines = [
            lines[j].strip()
            for j in range(ref_line_i + 1, ref_end)
            if lines[j].strip() and not _PAGE_COMMENT_RE.search(lines[j])
        ]
        marker_count = sum(1 for ln in ref_lines if _NUMBERED_LIST_RE.match(ln))
        items: list[str] = []
        if marker_count >= max(2, len(ref_lines) // 4):
            buf: list[str] = []

            def flush() -> None:
                if buf:
                    items.append(" ".join(s.strip() for s in buf if s.strip()))
                    buf.clear()

            for ln in ref_lines:
                m = _NUMBERED_LIST_RE.match(ln)
                if m:
                    flush()
                    buf.append(m.group(1))
                else:
                    buf.append(ln)
            flush()
        else:
            items = ref_lines

        for item in items:
            doi_match = _DOI_RE.search(item)
            references.append(
                {
                    "citation_text": item,
                    "doi": doi_match.group(1) if doi_match else None,
                }
            )

    return {
        "sections": sections,
        "tables": tables,
        "figures": figures,
        "references": references,
    }


def extract_paper_structure(
    file_stem: str,
    parsed_md_path: Path,
    paper_meta: dict,
    store: "KGStore",
) -> dict:
    md_text = Path(parsed_md_path).read_text(encoding="utf-8")
    parsed = _parse_markdown(md_text)
    sections = parsed["sections"]
    tables = parsed["tables"]
    figures = parsed["figures"]
    references = parsed["references"]

    counts = {
        "n_sections": len(sections),
        "n_tables": len(tables),
        "n_figures": len(figures),
        "n_references": len(references),
    }

    paper_props = {
        "paper_ref": paper_meta.get("paper_ref"),
        "doi": paper_meta.get("doi"),
        "title": paper_meta.get("title"),
        "year": paper_meta.get("year"),
        "source": paper_meta.get("source"),
        "full_md_path": str(parsed_md_path),
        "full_md_text": md_text,
        **counts,
    }

    store.execute_write(
        "MERGE (p:Paper {file_name: $file_name}) SET p += $props",
        file_name=file_stem,
        props=paper_props,
    )

    for idx, sec in enumerate(sections):
        sid = f"{file_stem}__sec_{idx:03d}"
        store.execute_write(
            """
            MATCH (p:Paper {file_name: $paper_file_name})
            MERGE (s:Section {section_id: $section_id})
            SET s += $props
            MERGE (p)-[:HAS_SECTION]->(s)
            """,
            paper_file_name=file_stem,
            section_id=sid,
            props={
                "section_id": sid,
                "paper_file_name": file_stem,
                "heading": sec["heading"],
                "level": sec["level"],
                "page": sec["page"],
                "text": sec["text"],
            },
        )

    for idx, tbl in enumerate(tables):
        tid = f"{file_stem}__tbl_{idx:03d}"
        store.execute_write(
            """
            MATCH (p:Paper {file_name: $paper_file_name})
            MERGE (t:PaperTable {table_id: $table_id})
            SET t += $props
            MERGE (p)-[:HAS_TABLE]->(t)
            """,
            paper_file_name=file_stem,
            table_id=tid,
            props={
                "table_id": tid,
                "paper_file_name": file_stem,
                "page": tbl["page"],
                "caption": tbl["caption"],
                "csv": tbl["csv"],
            },
        )

    for idx, fig in enumerate(figures):
        fid = f"{file_stem}__fig_{idx:03d}"
        store.execute_write(
            """
            MATCH (p:Paper {file_name: $paper_file_name})
            MERGE (f:Figure {figure_id: $figure_id})
            SET f += $props
            MERGE (p)-[:HAS_FIGURE]->(f)
            """,
            paper_file_name=file_stem,
            figure_id=fid,
            props={
                "figure_id": fid,
                "paper_file_name": file_stem,
                "page": fig["page"],
                "caption": fig["caption"],
            },
        )

    for idx, ref in enumerate(references):
        rid = f"{file_stem}__ref_{idx:03d}"
        store.execute_write(
            """
            MATCH (p:Paper {file_name: $paper_file_name})
            MERGE (r:Reference {ref_id: $ref_id})
            SET r += $props
            MERGE (p)-[:HAS_REFERENCE]->(r)
            """,
            paper_file_name=file_stem,
            ref_id=rid,
            props={
                "ref_id": rid,
                "paper_file_name": file_stem,
                "citation_text": ref["citation_text"],
                "doi": ref["doi"],
                "cited_paper_file_name": None,
            },
        )

    return {"file_stem": file_stem, **counts}


def _docling_json_to_markdown(json_path: Path) -> str:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    parts: list[str] = []
    last_page: int | None = None

    for sec in data.get("sections", []):
        page = sec.get("page")
        if page is not None and page != last_page:
            parts.append(f"<!-- page: {page} -->")
            last_page = page
        level = max(1, min(4, int(sec.get("level") or 1)))
        heading = (sec.get("heading") or "").strip()
        if heading:
            parts.append(f"{'#' * level} {heading}")
        body = (sec.get("text") or "").strip()
        if body:
            parts.append(body)
        parts.append("")

    for tbl in data.get("tables", []):
        page = tbl.get("page")
        if page is not None and page != last_page:
            parts.append(f"<!-- page: {page} -->")
            last_page = page
        md = (tbl.get("markdown") or "").strip()
        if md:
            parts.append(md)
            parts.append("")

    return "\n".join(parts)


def _paper_meta_for_stem(file_stem: str, doi: str | None) -> dict:
    year_match = re.search(r"(\d{4})$", file_stem)
    year = int(year_match.group(1)) if year_match else None
    author = file_stem.rsplit("_", 1)[0].replace("_", " ") if "_" in file_stem else file_stem
    paper_ref = f"{author} {year}" if year else file_stem
    return {
        "paper_ref": paper_ref,
        "doi": doi,
        "title": None,
        "year": year,
        "source": "provided",
    }


def _load_doi_index(xlsx_path: Path) -> dict[str, str]:
    if not xlsx_path.exists():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    fn_col = next((i for i, h in enumerate(header) if "file" in h), None)
    doi_col = next((i for i, h in enumerate(header) if "doi" in h), None)
    if fn_col is None or doi_col is None:
        return {}
    out: dict[str, str] = {}
    for row in rows[1:]:
        if not row:
            continue
        fn, doi = row[fn_col], row[doi_col]
        if not fn:
            continue
        stem = Path(str(fn)).stem
        if doi:
            out[stem] = str(doi).strip()
    return out


def extract_corpus_structure(
    parsed_dir: Path = Path("data/papers/parsed"),
    index_xlsx: Path = Path("data/papers/_index.xlsx"),
    store: "KGStore | None" = None,
) -> list[dict]:
    if store is None:
        from api.backends.kg_store import KGStore

        store = KGStore()
        store.bootstrap_schema()

    doi_index = _load_doi_index(index_xlsx)
    parsed_dir = Path(parsed_dir)
    md_files = sorted(parsed_dir.glob("*.md"))
    json_files = sorted(parsed_dir.glob("*.json"))
    seen: set[str] = set()
    summaries: list[dict] = []

    for md_path in md_files:
        stem = md_path.stem
        seen.add(stem)
        meta = _paper_meta_for_stem(stem, doi_index.get(stem))
        summaries.append(extract_paper_structure(stem, md_path, meta, store))

    for json_path in json_files:
        stem = json_path.stem
        if stem in seen:
            continue
        md_text = _docling_json_to_markdown(json_path)
        meta = _paper_meta_for_stem(stem, doi_index.get(stem))
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".md", delete=False, encoding="utf-8"
        ) as tf:
            tf.write(md_text)
            tmp_path = Path(tf.name)
        try:
            summaries.append(extract_paper_structure(stem, tmp_path, meta, store))
        finally:
            tmp_path.unlink(missing_ok=True)

    return summaries
